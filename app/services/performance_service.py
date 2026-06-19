
# v3 - with .pdf and excel full implimentations
"""
Student performance service - Complete professional implementation
with PDF/Excel export and comprehensive analytics.
"""

from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, and_, desc
from uuid import UUID
from typing import Dict, List, Any, Optional
from datetime import datetime, timedelta
from io import BytesIO

from sqlalchemy import func, and_, desc, or_
from app.models.enrollment import Enrollment
from app.models.scores import Score, ScoreColumn
from app.models.attendance import Attendance, AttendanceStatus
from app.models.modules import Module
from app.models.lesson import Lesson
from app.models.course import Course
from app.models.user import User
from fastapi import HTTPException, status

# PDF Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

# Excel Generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference


# ============================================================================
# GRADUATION CRITERIA
# ============================================================================
GRADUATION_CRITERIA = {
    "min_overall_average": 50.0,  # Minimum 50% overall
    "min_attendance_rate": 75.0,   # Minimum 75% attendance
    "min_assessments_completed": 0.8  # Must complete 80% of assessments
}


def get_student_performance(db: Session, student_id: UUID) -> Dict[str, Any]:
    """
    Get comprehensive performance data for a student across all enrolled courses.
    """
    enrollments = db.query(Enrollment).options(
        joinedload(Enrollment.course)
    ).filter(
        Enrollment.student_id == student_id
    ).all()
    
    if not enrollments:
        return {
            "summary": _get_empty_summary(),
            "courses": [],
            "attendance": _get_empty_attendance(),
            "attendance_details": [],
            "trends": [],
            "graduation_status": _get_graduation_status(None, None, 0, 0)
        }
    
    # Build course performance data
    courses_performance = []
    for enrollment in enrollments:
        course_data = _get_course_performance(db, enrollment)
        if course_data:
            courses_performance.append(course_data)
    
    # Calculate overall summary
    summary = _calculate_overall_summary(courses_performance)
    
    # Get attendance summary and details
    attendance = _get_attendance_summary(db, student_id)
    attendance_details = _get_attendance_details(db, student_id)
    
    # Get performance trends
    trends = _get_performance_trends(db, student_id, enrollments)
    
    # Calculate graduation eligibility
    total_assessments = sum(c['total_assessments'] for c in courses_performance)
    completed_assessments = sum(c['completed_assessments'] for c in courses_performance)
    graduation_status = _get_graduation_status(
        summary['overall_average'],
        attendance['attendance_rate'],
        completed_assessments,
        total_assessments
    )
    
    return {
        "summary": summary,
        "courses": courses_performance,
        "attendance": attendance,
        "attendance_details": attendance_details,
        "trends": trends,
        "graduation_status": graduation_status
    }


# def _get_course_performance(db: Session, enrollment: Enrollment) -> Dict[str, Any]:
#     """Calculate performance metrics for a single course with full assessment details."""
#     course = enrollment.course
    
#     # Get all scores with full details
#     scores = db.query(Score).options(
#         joinedload(Score.column)
#     ).filter(
#         Score.enrollment_id == enrollment.id
#     ).all()
    
#     # Categorize scores by scope with full titles
#     lesson_scores = []
#     module_scores = []
#     course_scores = []
    
#     for score in scores:
#         column = score.column
        
#         # Resolve scope title, lesson date, and scope ordering key
#         scope_title = None
#         lesson_date = None
#         scope_order = None  # for sorting within grouped view
        
#         if column.lesson_id:
#             lesson = db.query(Lesson).filter(Lesson.id == column.lesson_id).first()
#             scope_title = lesson.title if lesson else "Unknown Lesson"
#             # Expose the lesson's scheduled/actual date if the model carries one
#             if lesson and hasattr(lesson, 'date') and lesson.date:
#                 lesson_date = lesson.date.isoformat()
#             elif lesson and hasattr(lesson, 'date') and lesson.date:
#                 lesson_date = lesson.date.isoformat()
#             if lesson and hasattr(lesson, 'order'):
#                 scope_order = lesson.order
#         elif column.module_id:
#             module = db.query(Module).filter(Module.id == column.module_id).first()
#             scope_title = module.title if module else "Unknown Module"
#             if module and hasattr(module, 'order'):
#                 scope_order = module.order
#         elif column.course_id:
#             scope_title = course.title

#         # Resolve who recorded the score (teacher/staff name if available)
#         recorded_by = None
#         if hasattr(score, 'recorded_by_user') and score.recorded_by_user:
#             recorded_by = score.recorded_by_user.names or score.recorded_by_user.email
#         elif hasattr(score, 'recorded_by') and score.recorded_by:
#             # If it's a FK id rather than a relationship, skip — avoid extra query cost.
#             pass

#         score_data = {
#             "column_id": str(score.column_id),
#             "type": column.type.value,
#             "title": column.title,          # Assessment title (e.g., "Homework 1", "Quiz 2")
#             "scope_title": scope_title,     # Lesson / Module / Course title
#             "scope_order": scope_order,     # Ordering within its parent (for FE grouping)
#             "lesson_date": lesson_date,     # Scheduled/actual lesson date (if available)
#             "score": float(score.score) if score.score is not None else None,
#             "max_score": float(score.max_score),
#             "percentage": float(score.percentage) if score.percentage is not None else None,
#             "grade": score.grade if score.grade else "N/A",
#             "remarks": score.notes or "",
#             "recorded_date": score.recorded_date.isoformat() if score.recorded_date else None,
#             "recorded_by": recorded_by,     # Name/email of staff who recorded the score
#             "is_completed": score.score is not None,
#         }
        
#         if column.lesson_id:
#             score_data["lesson_id"] = str(column.lesson_id)
#             lesson_scores.append(score_data)
#         elif column.module_id:
#             score_data["module_id"] = str(column.module_id)
#             module_scores.append(score_data)
#         elif column.course_id:
#             score_data["course_id"] = str(column.course_id)
#             course_scores.append(score_data)
    
#     # Calculate metrics
#     total_scores = lesson_scores + module_scores + course_scores
#     completed_scores = [s for s in total_scores if s['is_completed']]
    
#     overall_average = 0.0
#     if completed_scores:
#         overall_average = round(
#             sum(s['percentage'] for s in completed_scores if s['percentage'] is not None) / len(completed_scores),
#             1
#         )
    
#     return {
#         "enrollment_id": str(enrollment.id),
#         "course": {
#             "id": str(course.id),
#             "title": course.title,
#             "code": course.code
#         },
#         "lesson_scores": lesson_scores,
#         "module_scores": module_scores,
#         "course_scores": course_scores,
#         "total_assessments": len(total_scores),
#         "completed_assessments": len(completed_scores),
#         "overall_average": overall_average,
#         "overall_grade": _calculate_grade_from_percentage(overall_average),
#         "enrolled_date": enrollment.enrolled_at.isoformat() if enrollment.enrolled_at else None
#     }


# v2
def _get_course_performance(db: Session, enrollment: Enrollment) -> Dict[str, Any]:
    """Calculate performance metrics for a single course with full assessment details."""
    course = enrollment.course

    # ========================================================================
    # STEP 1: Get all modules and lessons for this course (ensures ALL show)
    # ========================================================================
    modules = db.query(Module).filter(
        Module.course_id == course.id
    ).order_by(
        Module.order if hasattr(Module, 'order') else Module.id
    ).all()

    # ========================================================================
    # STEP 2: Get all score columns and scores for this enrollment
    # ========================================================================
    score_columns = db.query(ScoreColumn).outerjoin(
        Score,
        and_(Score.column_id == ScoreColumn.id, Score.enrollment_id == enrollment.id)
    ).options(
        joinedload(ScoreColumn.scores)
    ).filter(
        or_(
            ScoreColumn.course_id == course.id,
            ScoreColumn.module_id.in_([m.id for m in modules]),
            ScoreColumn.lesson_id.in_(
                db.query(Lesson.id).filter(Lesson.module_id.in_([m.id for m in modules]))
            )
        )
    ).all()

    # Build lookup: column_id -> score for this enrollment
    score_lookup: Dict[str, Any] = {}
    for col in score_columns:
        for s in col.scores:
            if s.enrollment_id == enrollment.id:
                score_lookup[str(col.id)] = s
                break

    # ========================================================================
    # STEP 3: Build lesson_scores - ALL lessons included, with their assessments
    # ========================================================================
    lesson_scores = []
    module_scores = []
    course_scores = []

    # Process each module and its lessons
    for module in modules:
        module_lessons = db.query(Lesson).filter(
            Lesson.module_id == module.id
        ).order_by(
            Lesson.order if hasattr(Lesson, 'order') else Lesson.id
        ).all()

        for lesson in module_lessons:
            # Find all ScoreColumns for this lesson
            lesson_columns = [col for col in score_columns if col.lesson_id == lesson.id]

            # Get lesson date
            lesson_date = None
            if hasattr(lesson, 'date') and lesson.date:
                lesson_date = lesson.date.isoformat()
            elif hasattr(lesson, 'scheduled_date') and lesson.scheduled_date:
                lesson_date = lesson.scheduled_date.isoformat()

            # Get module order for sorting
            module_order = None
            if hasattr(module, 'order'):
                module_order = module.order

            # Get lesson order for sorting
            lesson_order = None
            if hasattr(lesson, 'order'):
                lesson_order = lesson.order

            if lesson_columns:
                # Lesson has assessment columns - emit one score_data per column
                for column in lesson_columns:
                    score = score_lookup.get(str(column.id))

                    # Get recorded_by info
                    recorded_by = None
                    recorded_date = None
                    if score:
                        if score.recorded_date:
                            recorded_date = score.recorded_date.isoformat()
                        if hasattr(score, 'recorded_by') and score.recorded_by:
                            recorded_by = score.recorded_by.names or score.recorded_by.email or "Unknown"
                        elif hasattr(score, 'grader') and score.grader:
                            recorded_by = score.grader.names or score.grader.email or "Unknown"
                        elif hasattr(score, 'recorded_by_id') and score.recorded_by_id:
                            grader = db.query(User).filter(User.id == score.recorded_by_id).first()
                            recorded_by = grader.names if grader and grader.names else (grader.email if grader else "Unknown")
                        elif hasattr(score, 'grader_id') and score.grader_id:
                            grader = db.query(User).filter(User.id == score.grader_id).first()
                            recorded_by = grader.names if grader and grader.names else (grader.email if grader else "Unknown")

                    score_data = {
                        "column_id": str(column.id),
                        "type": column.type.value if column.type else "unknown",
                        "title": column.title or "Untitled Assessment",
                        "scope_title": lesson.title,
                        "module_name": module.title,
                        "module_order": module_order,
                        "lesson_order": lesson_order,
                        "score": float(score.score) if score and score.score is not None else None,
                        "max_score": float(column.max_score) if column.max_score else (float(score.max_score) if score and score.max_score else 0),
                        "percentage": float(score.percentage) if score and score.percentage is not None else None,
                        "grade": score.grade if score and score.grade else "N/A",
                        "remarks": (score.notes if score and score.notes else "") or column.description or "",
                        "recorded_date": recorded_date,
                        "recorded_by": recorded_by,
                        "lesson_date": lesson_date,
                        "is_completed": score is not None and score.score is not None
                    }
                    lesson_scores.append(score_data)
            else:
                # Lesson has NO assessment columns - emit a placeholder row
                # so the lesson still appears in the table
                placeholder = {
                    "column_id": f"placeholder-{lesson.id}",
                    "type": "none",
                    "title": "No assessments",
                    "scope_title": lesson.title,
                    "module_name": module.title,
                    "module_order": module_order,
                    "lesson_order": lesson_order,
                    "score": None,
                    "max_score": 0,
                    "percentage": None,
                    "grade": "N/A",
                    "remarks": "",
                    "recorded_date": None,
                    "recorded_by": None,
                    "lesson_date": lesson_date,
                    "is_completed": False
                }
                lesson_scores.append(placeholder)

    # ========================================================================
    # STEP 4: Process module-level and course-level assessments (unchanged logic)
    # ========================================================================
    for column in score_columns:
        if column.module_id and not column.lesson_id:
            score = score_lookup.get(str(column.id))
            module = db.query(Module).filter(Module.id == column.module_id).first()

            recorded_by = None
            recorded_date = None
            if score:
                if score.recorded_date:
                    recorded_date = score.recorded_date.isoformat()
                if hasattr(score, 'recorded_by') and score.recorded_by:
                    recorded_by = score.recorded_by.names or score.recorded_by.email or "Unknown"
                elif hasattr(score, 'grader') and score.grader:
                    recorded_by = score.grader.names or score.grader.email or "Unknown"
                elif hasattr(score, 'recorded_by_id') and score.recorded_by_id:
                    grader = db.query(User).filter(User.id == score.recorded_by_id).first()
                    recorded_by = grader.names if grader and grader.names else (grader.email if grader else "Unknown")
                elif hasattr(score, 'grader_id') and score.grader_id:
                    grader = db.query(User).filter(User.id == score.grader_id).first()
                    recorded_by = grader.names if grader and grader.names else (grader.email if grader else "Unknown")

            score_data = {
                "column_id": str(column.id),
                "type": column.type.value if column.type else "unknown",
                "title": column.title or "Untitled Assessment",
                "scope_title": module.title if module else "Unknown Module",
                "module_name": module.title if module else "Unknown",
                "module_order": module.order if module and hasattr(module, 'order') else None,
                "lesson_order": None,
                "score": float(score.score) if score and score.score is not None else None,
                "max_score": float(column.max_score) if column.max_score else (float(score.max_score) if score and score.max_score else 0),
                "percentage": float(score.percentage) if score and score.percentage is not None else None,
                "grade": score.grade if score and score.grade else "N/A",
                "remarks": (score.notes if score and score.notes else "") or column.description or "",
                "recorded_date": recorded_date,
                "recorded_by": recorded_by,
                "lesson_date": None,
                "is_completed": score is not None and score.score is not None
            }
            module_scores.append(score_data)

        elif column.course_id and not column.module_id and not column.lesson_id:
            score = score_lookup.get(str(column.id))

            recorded_by = None
            recorded_date = None
            if score:
                if score.recorded_date:
                    recorded_date = score.recorded_date.isoformat()
                if hasattr(score, 'recorded_by') and score.recorded_by:
                    recorded_by = score.recorded_by.names or score.recorded_by.email or "Unknown"
                elif hasattr(score, 'grader') and score.grader:
                    recorded_by = score.grader.names or score.grader.email or "Unknown"
                elif hasattr(score, 'recorded_by_id') and score.recorded_by_id:
                    grader = db.query(User).filter(User.id == score.recorded_by_id).first()
                    recorded_by = grader.names if grader and grader.names else (grader.email if grader else "Unknown")
                elif hasattr(score, 'grader_id') and score.grader_id:
                    grader = db.query(User).filter(User.id == score.grader_id).first()
                    recorded_by = grader.names if grader and grader.names else (grader.email if grader else "Unknown")

            score_data = {
                "column_id": str(column.id),
                "type": column.type.value if column.type else "unknown",
                "title": column.title or "Untitled Assessment",
                "scope_title": course.title,
                "module_name": None,
                "module_order": None,
                "lesson_order": None,
                "score": float(score.score) if score and score.score is not None else None,
                "max_score": float(column.max_score) if column.max_score else (float(score.max_score) if score and score.max_score else 0),
                "percentage": float(score.percentage) if score and score.percentage is not None else None,
                "grade": score.grade if score and score.grade else "N/A",
                "remarks": (score.notes if score and score.notes else "") or column.description or "",
                "recorded_date": recorded_date,
                "recorded_by": recorded_by,
                "lesson_date": None,
                "is_completed": score is not None and score.score is not None
            }
            course_scores.append(score_data)

    # ========================================================================
    # STEP 5: Sort and calculate metrics
    # ========================================================================
    lesson_scores.sort(key=lambda x: (x.get('module_order') or 0, x.get('lesson_order') or 0, x.get('column_id') or ''))
    module_scores.sort(key=lambda x: x.get('module_order') or 0)

    total_scores = lesson_scores + module_scores + course_scores
    completed_scores = [s for s in total_scores if s['is_completed']]

    overall_average = 0.0
    if completed_scores:
        overall_average = round(
            sum(s['percentage'] for s in completed_scores if s['percentage'] is not None) / len(completed_scores),
            1
        )

    return {
        "enrollment_id": str(enrollment.id),
        "course": {
            "id": str(course.id),
            "title": course.title,
            "code": course.code
        },
        "lesson_scores": lesson_scores,
        "module_scores": module_scores,
        "course_scores": course_scores,
        "total_assessments": len(total_scores),
        "completed_assessments": len(completed_scores),
        "overall_average": overall_average,
        "overall_grade": _calculate_grade_from_percentage(overall_average),
        "enrolled_date": enrollment.enrolled_at.isoformat() if enrollment.enrolled_at else None
    }

def _get_attendance_details(db: Session, student_id: UUID) -> List[Dict[str, Any]]:
    """Get detailed attendance records for student."""
    records = db.query(Attendance).options(
        joinedload(Attendance.lesson).joinedload(Lesson.module).joinedload(Module.course)
    ).filter(
        Attendance.student_id == student_id
    ).order_by(
        desc(Attendance.date)
    ).all()
    
    details = []
    for record in records:
        lesson = record.lesson
        module = lesson.module if lesson else None
        course = module.course if module else None
        
        # Resolve lesson date (scheduled or actual)
        lesson_date = None
        if lesson:
            if hasattr(lesson, "date") and lesson.date:
                lesson_date = lesson.date.isoformat()
            elif hasattr(lesson, "date") and lesson.date:
                lesson_date = lesson.date.isoformat()

        # Resolve who recorded this attendance entry
        recorded_by = None
        if hasattr(record, "created_by") and record.created_by:
            recorded_by = record.created_by.names or record.created_by.email

        details.append({
            "id": str(record.id),
            "date": record.date.isoformat(),
            "status": record.status.value,
            "lesson_title": lesson.title if lesson else "Unknown",
            "lesson_date": lesson_date,
            "module_title": module.title if module else "Unknown",
            "course_title": course.title if course else "Unknown",
            "course_code": course.code if course else "N/A",
            "remarks": record.notes or "",
            "recorded_by": recorded_by,
        })
    
    return details


def _get_graduation_status(
    overall_average: Optional[float],
    attendance_rate: Optional[float],
    completed: int,
    total: int
) -> Dict[str, Any]:
    """Determine if student qualifies for graduation/certification."""
    if overall_average is None or attendance_rate is None:
        return {
            "qualified": False,
            "status": "incomplete",
            "message": "Insufficient data to determine graduation eligibility",
            "criteria_met": {},
            "recommendations": ["Complete more assessments to qualify"]
        }
    
    completion_rate = (completed / total * 100) if total > 0 else 0
    
    criteria_met = {
        "academic_performance": overall_average >= GRADUATION_CRITERIA["min_overall_average"],
        "attendance": attendance_rate >= GRADUATION_CRITERIA["min_attendance_rate"],
        "completion": completion_rate >= (GRADUATION_CRITERIA["min_assessments_completed"] * 100)
    }
    
    all_qualified = all(criteria_met.values())
    
    # Generate recommendations
    recommendations = []
    if not criteria_met["academic_performance"]:
        needed = GRADUATION_CRITERIA["min_overall_average"] - overall_average
        recommendations.append(f"Improve grades by {needed:.1f}% to meet minimum requirement")
    
    if not criteria_met["attendance"]:
        needed = GRADUATION_CRITERIA["min_attendance_rate"] - attendance_rate
        recommendations.append(f"Improve attendance by {needed:.1f}% to meet requirement")
    
    if not criteria_met["completion"]:
        needed = int(total * GRADUATION_CRITERIA["min_assessments_completed"]) - completed
        recommendations.append(f"Complete {needed} more assessments")
    
    if all_qualified:
        status = "qualified"
        message = "Congratulations! You meet all requirements for graduation/certification."
        recommendations = ["Maintain your excellent performance"]
    else:
        status = "not_qualified"
        message = "You do not currently meet graduation requirements."
    
    return {
        "qualified": all_qualified,
        "status": status,
        "message": message,
        "criteria": {
            "min_average": GRADUATION_CRITERIA["min_overall_average"],
            "min_attendance": GRADUATION_CRITERIA["min_attendance_rate"],
            "min_completion": GRADUATION_CRITERIA["min_assessments_completed"] * 100
        },
        "current": {
            "average": overall_average,
            "attendance": attendance_rate,
            "completion": completion_rate
        },
        "criteria_met": criteria_met,
        "recommendations": recommendations
    }


# ============================================================================
# PDF EXPORT
# ============================================================================

def export_performance_pdf(db: Session, student_id: UUID) -> bytes:
    """Generate professional PDF performance report."""
    from app.models.user import User
    
    performance = get_student_performance(db, student_id)
    student = db.query(User).filter(User.id == student_id).first()
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72,
                           topMargin=72, bottomMargin=18)
    
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Center', alignment=TA_CENTER))
    styles.add(ParagraphStyle(name='Right', alignment=TA_RIGHT))
    
    story = []
    
    # Title
    story.append(Paragraph("<b>ACADEMIC PERFORMANCE REPORT</b>", styles['Title']))
    story.append(Spacer(1, 12))
    
    # Student Info
    story.append(Paragraph(f"<b>Student:</b> {student.names if student else 'Unknown'}", styles['Normal']))
    story.append(Paragraph(f"<b>Email:</b> {student.email if student else 'N/A'}", styles['Normal']))
    story.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Summary
    summary = performance['summary']
    story.append(Paragraph("<b>OVERALL SUMMARY</b>", styles['Heading2']))
    summary_data = [
        ['Metric', 'Value'],
        ['Total Courses', str(summary['total_courses'])],
        ['Total Assessments', str(summary['total_assessments'])],
        ['Overall Average', f"{summary['overall_average']}%"],
        ['Overall Grade', summary['overall_grade']]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Graduation Status
    grad = performance['graduation_status']
    story.append(Paragraph("<b>GRADUATION STATUS</b>", styles['Heading2']))
    grad_color = colors.green if grad['qualified'] else colors.red
    story.append(Paragraph(f"<b>Status:</b> <font color='{grad_color}'>{grad['message']}</font>", styles['Normal']))
    story.append(Spacer(1, 10))
    
    # Course Details
    story.append(Paragraph("<b>COURSE PERFORMANCE</b>", styles['Heading2']))
    for course in performance['courses']:
        story.append(Paragraph(f"<b>{course['course']['code']}: {course['course']['title']}</b>", styles['Heading3']))
        story.append(Paragraph(f"Grade: {course['overall_grade']} ({course['overall_average']}%)", styles['Normal']))
        story.append(Paragraph(f"Assessments: {course['completed_assessments']}/{course['total_assessments']}", styles['Normal']))
        story.append(Spacer(1, 10))
    
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================================================
# EXCEL EXPORT
# ============================================================================

def export_performance_excel(db: Session, student_id: UUID) -> bytes:
    """Generate professional Excel performance report."""
    from app.models.user import User
    
    performance = get_student_performance(db, student_id)
    student = db.query(User).filter(User.id == student_id).first()
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Summary Sheet
    ws_summary = wb.create_sheet("Summary")
    _create_summary_sheet(ws_summary, student, performance)
    
    # Courses Sheet
    ws_courses = wb.create_sheet("Courses")
    _create_courses_sheet(ws_courses, performance['courses'])
    
    # Attendance Sheet
    ws_attendance = wb.create_sheet("Attendance")
    _create_attendance_sheet(ws_attendance, performance['attendance_details'])
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _create_summary_sheet(ws, student, performance):
    """Create summary worksheet."""
    # Title
    ws['A1'] = 'ACADEMIC PERFORMANCE REPORT'
    ws['A1'].font = Font(bold=True, size=16)
    
    # Student info
    ws['A3'] = 'Student Name:'
    ws['B3'] = student.names if student else 'Unknown'
    ws['A4'] = 'Email:'
    ws['B4'] = student.email if student else 'N/A'
    ws['A5'] = 'Report Date:'
    ws['B5'] = datetime.now().strftime('%B %d, %Y')
    
    # Summary
    summary = performance['summary']
    ws['A7'] = 'OVERALL SUMMARY'
    ws['A7'].font = Font(bold=True, size=14)
    
    headers = ['Metric', 'Value']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    data = [
        ('Total Courses', summary['total_courses']),
        ('Total Assessments', summary['total_assessments']),
        ('Overall Average', f"{summary['overall_average']}%"),
        ('Overall Grade', summary['overall_grade'])
    ]
    
    for row, (metric, value) in enumerate(data, 9):
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=value)
    
    # Graduation Status
    grad = performance['graduation_status']
    ws['A14'] = 'GRADUATION STATUS'
    ws['A14'].font = Font(bold=True, size=14)
    ws['A15'] = 'Status:'
    ws['B15'] = grad['message']
    ws['B15'].font = Font(color='008000' if grad['qualified'] else 'FF0000', bold=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30


def _create_courses_sheet(ws, courses):
    """Create courses worksheet."""
    ws['A1'] = 'COURSE PERFORMANCE DETAILS'
    ws['A1'].font = Font(bold=True, size=14)
    
    headers = ['Course Code', 'Course Title', 'Grade', 'Average (%)', 'Assessments', 'Completed']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    for row, course in enumerate(courses, 4):
        ws.cell(row=row, column=1, value=course['course']['code'])
        ws.cell(row=row, column=2, value=course['course']['title'])
        ws.cell(row=row, column=3, value=course['overall_grade'])
        ws.cell(row=row, column=4, value=course['overall_average'])
        ws.cell(row=row, column=5, value=course['total_assessments'])
        ws.cell(row=row, column=6, value=course['completed_assessments'])
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15


def _create_attendance_sheet(ws, attendance_details):
    """Create attendance worksheet."""
    ws['A1'] = 'ATTENDANCE RECORD'
    ws['A1'].font = Font(bold=True, size=14)
    
    headers = ['Date', 'Status', 'Course', 'Module', 'Lesson', 'Remarks']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    for row, record in enumerate(attendance_details, 4):
        ws.cell(row=row, column=1, value=record['date'])
        
        status_cell = ws.cell(row=row, column=2, value=record['status'].upper())
        if record['status'] == 'present':
            status_cell.font = Font(color='008000')
        elif record['status'] == 'absent':
            status_cell.font = Font(color='FF0000')
        else:
            status_cell.font = Font(color='FFA500')
        
        ws.cell(row=row, column=3, value=f"{record['course_code']} - {record['course_title']}")
        ws.cell(row=row, column=4, value=record['module_title'])
        ws.cell(row=row, column=5, value=record['lesson_title'])
        ws.cell(row=row, column=6, value=record['remarks'])
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 30


# ============================================================================
# HELPER FUNCTIONS (CONTINUED)
# ============================================================================

def _calculate_grade_from_percentage(percentage: float) -> str:
    """Convert percentage to letter grade."""
    if percentage >= 90:
        return "A+"
    elif percentage >= 80:
        return "A"
    elif percentage >= 75:
        return "B+"
    elif percentage >= 70:
        return "B"
    elif percentage >= 65:
        return "C+"
    elif percentage >= 60:
        return "C"
    elif percentage >= 55:
        return "D+"
    elif percentage >= 50:
        return "D"
    else:
        return "F"


def _calculate_overall_summary(courses: List[Dict]) -> Dict[str, Any]:
    """Calculate overall summary across all courses."""
    if not courses:
        return _get_empty_summary()
    
    total_courses = len(courses)
    total_assessments = sum(c['total_assessments'] for c in courses)
    overall_average = round(
        sum(c['overall_average'] for c in courses) / total_courses,
        1
    ) if total_courses > 0 else 0.0
    
    grades = [c['overall_grade'] for c in courses]
    grade_distribution = {
        "A+": grades.count("A+"),
        "A": grades.count("A"),
        "B+": grades.count("B+"),
        "B": grades.count("B"),
        "C+": grades.count("C+"),
        "C": grades.count("C"),
        "D+": grades.count("D+"),
        "D": grades.count("D"),
        "F": grades.count("F")
    }
    
    return {
        "total_courses": total_courses,
        "total_assessments": total_assessments,
        "overall_average": overall_average,
        "overall_grade": _calculate_grade_from_percentage(overall_average),
        "grade_distribution": grade_distribution
    }


def _get_attendance_summary(db: Session, student_id: UUID) -> Dict[str, Any]:
    """Get attendance summary for student."""
    total = db.query(func.count(Attendance.id)).filter(
        Attendance.student_id == student_id
    ).scalar() or 0
    
    present = db.query(func.count(Attendance.id)).filter(
        and_(
            Attendance.student_id == student_id,
            Attendance.status == AttendanceStatus.PRESENT
        )
    ).scalar() or 0
    
    absent = db.query(func.count(Attendance.id)).filter(
        and_(
            Attendance.student_id == student_id,
            Attendance.status == AttendanceStatus.ABSENT
        )
    ).scalar() or 0
    
    late = db.query(func.count(Attendance.id)).filter(
        and_(
            Attendance.student_id == student_id,
            Attendance.status == AttendanceStatus.LATE
        )
    ).scalar() or 0
    
    attendance_rate = round((present / total * 100), 1) if total > 0 else 0.0
    
    return {
        "total": total,
        "present": present,
        "absent": absent,
        "late": late,
        "attendance_rate": attendance_rate
    }


def _get_performance_trends(
    db: Session, 
    student_id: UUID,
    enrollments: List[Enrollment]
) -> List[Dict]:
    """Get performance trends over time."""
    enrollment_ids = [e.id for e in enrollments]
    
    if not enrollment_ids:
        return []
    
    six_months_ago = datetime.utcnow() - timedelta(days=180)
    
    scores = db.query(
        func.date_trunc('month', Score.recorded_date).label('month'),
        func.avg(Score.percentage).label('avg_percentage')
    ).filter(
        and_(
            Score.enrollment_id.in_(enrollment_ids),
            Score.recorded_date >= six_months_ago,
            Score.recorded_date.isnot(None),
            Score.percentage.isnot(None)
        )
    ).group_by(
        func.date_trunc('month', Score.recorded_date)
    ).order_by(
        func.date_trunc('month', Score.recorded_date)
    ).all()
    
    return [
        {
            "month": score.month.strftime("%B %Y") if score.month else "Unknown",
            "average": round(score.avg_percentage, 1) if score.avg_percentage else 0
        }
        for score in scores
    ]


def _get_empty_summary() -> Dict[str, Any]:
    """Return empty summary structure."""
    return {
        "total_courses": 0,
        "total_assessments": 0,
        "overall_average": 0.0,
        "overall_grade": "N/A",
        "grade_distribution": {
            "A+": 0, "A": 0, "B+": 0, "B": 0, 
            "C+": 0, "C": 0, "D+": 0, "D": 0, "F": 0
        }
    }


def _get_empty_attendance() -> Dict[str, Any]:
    """Return empty attendance structure."""
    return {
        "total": 0,
        "present": 0,
        "absent": 0,
        "late": 0,
        "attendance_rate": 0.0
    }

